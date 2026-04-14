import openpyxl
import json
import os
from datetime import datetime

def custom_serializer(obj):
    """Custom serializer for non-serializable data types."""
    if isinstance(obj, datetime):
        return obj.isoformat()
    return str(obj)

def read_excel_structure(file_path):
    """
    Reads the structure of an Excel file and outputs its sheet names, columns, and example data.
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file '{file_path}' does not exist.")

        workbook = openpyxl.load_workbook(file_path)
        structure = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            structure[sheet_name] = {
                "columns": [cell.value for cell in sheet[1] if cell.value is not None],
                "example_data": [
                    [custom_serializer(cell.value) for cell in row] for row in sheet.iter_rows(min_row=2)
                ],
            }

        output_path = os.path.join(os.getcwd(), "excel_structure.json")
        with open(output_path, "w", encoding="utf-8") as json_file:
            json.dump(structure, json_file, ensure_ascii=False, indent=4)

        print(f"Excel structure has been saved to '{output_path}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    # Use absolute path for the Excel file
    excel_file_path = os.path.abspath("cubiX-SGMW - System Requirements_CST.xlsx")
    read_excel_structure(excel_file_path)