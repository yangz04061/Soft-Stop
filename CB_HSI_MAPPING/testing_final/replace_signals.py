import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
from copy import copy
from datetime import datetime

# 文件路径
source_file = r"cubiX-SGMW - System Requirements.xlsx"
mapping_file = r"SGMW_external_interface_to_Kun_20260407.xlsx"
backup_file = f"cubiX-SGMW - System Requirements_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# 步骤1：读取映射文件，建立信号映射
print("=" * 80)
print("步骤1：读取映射文件中的信号对应关系...")
print("=" * 80)

mapping_wb = openpyxl.load_workbook(mapping_file, data_only=True)
signal_mapping = {}  # SGMW Ext. Interface Name -> Match_HSI_Signal

for sheet_name in mapping_wb.sheetnames:
    print(f"\n处理Sheet: {sheet_name}")
    sheet = mapping_wb[sheet_name]
    
    # 查找列的标题
    header_row = None
    sgmw_col = None
    match_hsi_col = None
    
    # 查找包含"SGMW Ext. Interface Name"和"Match_HSI_Signal"的列
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            if cell.value and "SGMW Ext. Interface Name" in str(cell.value):
                sgmw_col = idx + 1
            if cell.value and "Match_HSI_Signal" in str(cell.value):
                match_hsi_col = idx + 1
        if sgmw_col is not None and match_hsi_col is not None:
            header_row = next((i+1 for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10)) 
                              if any(cell.value and "SGMW Ext. Interface Name" in str(cell.value) for cell in row)), None)
            break
    
    if sgmw_col is None or match_hsi_col is None:
        print(f"  警告：在Sheet {sheet_name}中未找到需要的列")
        print(f"    SGMW Ext. Interface Name列: {sgmw_col}, Match_HSI_Signal列: {match_hsi_col}")
        continue
    
    # 读取数据
    for row in sheet.iter_rows(min_row=header_row+1 if header_row else 2):
        sgmw_signal = row[sgmw_col-1].value
        match_signal = row[match_hsi_col-1].value
        
        if sgmw_signal and match_signal:
            signal_mapping[str(sgmw_signal).strip()] = str(match_signal).strip()
            print(f"  映射: {sgmw_signal} -> {match_signal}")

print(f"\n总共建立了 {len(signal_mapping)} 条映射关系")
print("\n映射关系详情:")
for k, v in sorted(signal_mapping.items()):
    print(f"  {k:50s} -> {v}")

# 步骤2：打开System Requirements文件
print("\n" + "=" * 80)
print("步骤2：保存原文件备份...")
print("=" * 80)

from shutil import copy as shutil_copy
shutil_copy(source_file, backup_file)
print(f"备份文件已保存: {backup_file}")

# 步骤3：读取和修改工作文件
print("\n" + "=" * 80)
print("步骤3：在工作文件中进行替换...")
print("=" * 80)

wb = openpyxl.load_workbook(source_file)
replacement_log = []

# 检查是否有{}中的内容，进行替换
def replace_signals_in_text(text, mapping, sheet_name, row_num, col_num):
    """在文本中查找{}包含的信号，进行替换"""
    if not text or not isinstance(text, str):
        return text, []
    
    replacements = []
    
    # 查找所有{}中的内容
    pattern = r'\{([^}]+)\}'
    matches = re.findall(pattern, text)
    
    result = text
    for match in matches:
        signal_name = match.strip()
        if signal_name in mapping:
            old_text = f"{{{signal_name}}}"
            new_text = f"{{{mapping[signal_name]}}}"
            result = result.replace(old_text, new_text)
            replacements.append({
                'sheet': sheet_name,
                'row': row_num,
                'col': col_num,
                'old_signal': signal_name,
                'new_signal': mapping[signal_name],
                'status': 'SUCCESS'
            })
        else:
            replacements.append({
                'sheet': sheet_name,
                'row': row_num,
                'col': col_num,
                'old_signal': signal_name,
                'new_signal': None,
                'status': 'NOT_FOUND'
            })
    
    return result, replacements

# 处理所有Sheet
for sheet_name in wb.sheetnames:
    print(f"\n处理Sheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    sheet_replacements = []
    
    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '{' in cell.value:
                new_text, replacements = replace_signals_in_text(
                    cell.value, signal_mapping, sheet_name, cell.row, cell.column
                )
                
                if replacements:
                    cell.value = new_text
                    sheet_replacements.extend(replacements)
    
    if sheet_replacements:
        print(f"  该Sheet共进行了 {len(sheet_replacements)} 项替换操作")
        replacement_log.extend(sheet_replacements)

print(f"\n总共进行了 {len(replacement_log)} 项替换操作")

# 步骤4：保存修改后的文件
print("\n" + "=" * 80)
print("步骤4：保存修改后的工作文件...")
print("=" * 80)

wb.save(source_file)
print(f"工作文件已更新: {source_file}")

# 步骤5：生成替换分析报告
print("\n" + "=" * 80)
print("步骤5：生成替换分析报告...")
print("=" * 80)

# 按状态分类
successful = [r for r in replacement_log if r['status'] == 'SUCCESS']
not_found = [r for r in replacement_log if r['status'] == 'NOT_FOUND']

print(f"\n成功替换: {len(successful)} 项")
print(f"未找到映射: {len(not_found)} 项")

if successful:
    print("\n" + "=" * 80)
    print("成功替换的信号:")
    print("=" * 80)
    for r in successful:
        print(f"  Sheet: {r['sheet']:20s} | 行: {r['row']:5d} | 列: {r['col']:5d}")
        print(f"    {r['old_signal']:50s} -> {r['new_signal']}")

if not_found:
    print("\n" + "=" * 80)
    print("未找到映射的信号 (原文件中存在但映射文件中不存在):")
    print("=" * 80)
    for r in not_found:
        print(f"  Sheet: {r['sheet']:20s} | 行: {r['row']:5d} | 列: {r['col']:5d}")
        print(f"    {r['old_signal']:50s} (未找到对应的Match_HSI_Signal)")

# 生成详细报告文件
report_file = f"替换分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
with open(report_file, 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("信号替换分析报告\n")
    f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("=" * 80 + "\n\n")
    
    f.write(f"源文件: {source_file}\n")
    f.write(f"映射文件: {mapping_file}\n")
    f.write(f"备份文件: {backup_file}\n\n")
    
    f.write(f"映射关系总数: {len(signal_mapping)}\n")
    f.write(f"成功替换数: {len(successful)}\n")
    f.write(f"未找到映射数: {len(not_found)}\n\n")
    
    f.write("=" * 80 + "\n")
    f.write("映射关系详情:\n")
    f.write("=" * 80 + "\n")
    for k, v in sorted(signal_mapping.items()):
        f.write(f"{k:50s} -> {v}\n")
    
    if successful:
        f.write("\n" + "=" * 80 + "\n")
        f.write("成功替换的信号:\n")
        f.write("=" * 80 + "\n")
        for r in successful:
            f.write(f"Sheet: {r['sheet']:20s} 行号: {r['row']:5d} 列号: {r['col']:5d}\n")
            f.write(f"  {r['old_signal']:50s} -> {r['new_signal']}\n")
    
    if not_found:
        f.write("\n" + "=" * 80 + "\n")
        f.write("未找到映射的信号:\n")
        f.write("=" * 80 + "\n")
        for r in not_found:
            f.write(f"Sheet: {r['sheet']:20s} 行号: {r['row']:5d} 列号: {r['col']:5d}\n")
            f.write(f"  {r['old_signal']:50s} (未找到对应的Match_HSI_Signal)\n")

print(f"\n详细报告已保存: {report_file}")

print("\n" + "=" * 80)
print("任务完成!")
print("=" * 80)
