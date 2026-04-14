import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
from copy import copy
from datetime import datetime

# 文件路径
source_file = r"cubiX-SGMW - System Requirements.xlsx"
mapping_file = r"SGMW_external_interface_to_Kun_20260407.xlsx"
backup_file = f"cubiX-SGMW - System Requirements_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# 步骤1：读取映射文件，建立信号映射
print("=" * 80)
print("步骤1：读取映射文件中的信号对应关系...")
print("=" * 80)

mapping_wb = openpyxl.load_workbook(mapping_file, data_only=True)
signal_mapping = {}  # SGMW Ext. Interface Name -> Match_HSI_Signal

for sheet_name in mapping_wb.sheetnames:
    print(f"\n处理Sheet: {sheet_name}")
    sheet = mapping_wb[sheet_name]
    
    # 查找列的标题
    header_row = None
    sgmw_col = None
    match_hsi_col = None
    
    # 查找包含"SGMW Ext. Interface Name"和"Match_HSI_Signal"的列
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for idx, cell in enumerate(row):
            if cell.value and "SGMW Ext. Interface Name" in str(cell.value):
                sgmw_col = idx + 1
            if cell.value and "Match_HSI_Signal" in str(cell.value):
                match_hsi_col = idx + 1
        if sgmw_col is not None and match_hsi_col is not None:
            header_row = next((i+1 for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10)) 
                              if any(cell.value and "SGMW Ext. Interface Name" in str(cell.value) for cell in row)), None)
            break
    
    if sgmw_col is None or match_hsi_col is None:
        print(f"  警告：在Sheet {sheet_name}中未找到需要的列")
        print(f"    SGMW Ext. Interface Name列: {sgmw_col}, Match_HSI_Signal列: {match_hsi_col}")
        continue
    
    # 读取数据
    for row in sheet.iter_rows(min_row=header_row+1 if header_row else 2):
        sgmw_signal = row[sgmw_col-1].value
        match_signal = row[match_hsi_col-1].value
        
        if sgmw_signal and match_signal:
            sgmw_signal_str = str(sgmw_signal).strip()
            match_signal_str = str(match_signal).strip()
            signal_mapping[sgmw_signal_str] = match_signal_str
            # 也添加去掉~和其他干扰符号的版本
            signal_mapping_clean = sgmw_signal_str.replace("~", "").strip()
            if signal_mapping_clean != sgmw_signal_str:
                signal_mapping[signal_mapping_clean] = match_signal_str

print(f"\n总共建立了 {len(signal_mapping)} 条映射关系")

# 步骤2：保存备份
print("\n" + "=" * 80)
print("步骤2：保存原文件备份...")
print("=" * 80)

from shutil import copy as shutil_copy
shutil_copy(source_file, backup_file)
print(f"备份文件已保存: {backup_file}")

# 步骤3：改进的替换函数
print("\n" + "=" * 80)
print("步骤3：在工作文件中进行替换...")
print("=" * 80)

wb = openpyxl.load_workbook(source_file)
replacement_log = []

def replace_signals_in_text(text, mapping, sheet_name, row_num, col_num):
    """在文本中查找{}包含的信号，进行替换"""
    if not text or not isinstance(text, str):
        return text, []
    
    replacements = []
    
    # 查找所有{}中的内容
    pattern = r'\{([^}]+)\}'
    matches = re.findall(pattern, text)
    
    result = text
    for match in matches:
        signal_name_raw = match.strip()
        
        # 尝试多种方式来匹配信号名
        found = False
        matched_signal = None
        
        # 1. 直接匹配
        if signal_name_raw in mapping:
            matched_signal = mapping[signal_name_raw]
            found = True
        else:
            # 2. 尝试移除特殊字符后匹配
            signal_name_clean = signal_name_raw.replace("~", "").replace(" ", "").strip()
            if signal_name_clean in mapping:
                matched_signal = mapping[signal_name_clean]
                found = True
            else:
                # 3. 尝试在映射中查找相似的（去掉~后）
                for map_key, map_value in mapping.items():
                    map_key_clean = map_key.replace("~", "").replace(" ", "").strip()
                    if map_key_clean == signal_name_clean and signal_name_clean:
                        matched_signal = map_value
                        found = True
                        break
        
        if found and matched_signal:
            old_text = f"{{{signal_name_raw}}}"
            new_text = f"{{{matched_signal}}}"
            result = result.replace(old_text, new_text)
            replacements.append({
                'sheet': sheet_name,
                'row': row_num,
                'col': col_num,
                'old_signal': signal_name_raw,
                'new_signal': matched_signal,
                'status': 'SUCCESS'
            })
        else:
            replacements.append({
                'sheet': sheet_name,
                'row': row_num,
                'col': col_num,
                'old_signal': signal_name_raw,
                'new_signal': None,
                'status': 'NOT_FOUND'
            })
    
    return result, replacements

# 处理所有Sheet
for sheet_name in wb.sheetnames:
    print(f"\n处理Sheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    sheet_replacements = []
    
    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '{' in cell.value and '}' in cell.value:
                new_text, replacements = replace_signals_in_text(
                    cell.value, signal_mapping, sheet_name, cell.row, cell.column
                )
                
                if any(r['status'] == 'SUCCESS' for r in replacements):
                    cell.value = new_text
                    sheet_replacements.extend(replacements)
                    
                if any(r['status'] != 'SUCCESS' for r in replacements):
                    sheet_replacements.extend([r for r in replacements if r['status'] != 'SUCCESS'])
    
    if sheet_replacements:
        success_count = len([r for r in sheet_replacements if r['status'] == 'SUCCESS'])
        print(f"  该Sheet共进行了 {success_count} 项成功替换操作")
        replacement_log.extend(sheet_replacements)

successful = [r for r in replacement_log if r['status'] == 'SUCCESS']
not_found = [r for r in replacement_log if r['status'] == 'NOT_FOUND']

print(f"\n总共成功替换了 {len(successful)} 项")
print(f"未找到映射有 {len(not_found)} 项")

# 步骤4：保存修改后的文件
print("\n" + "=" * 80)
print("步骤4：保存修改后的工作文件...")
print("=" * 80)

wb.save(source_file)
print(f"工作文件已更新: {source_file}")

# 步骤5：生成替换分析报告
print("\n" + "=" * 80)
print("步骤5：生成详细替换分析报告...")
print("=" * 80)

# 按状态分类
print(f"\n成功替换: {len(successful)} 项")
print(f"未找到映射: {len(not_found)} 项")

# 生成详细报告文件
report_file = f"替换分析报告_详细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
with open(report_file, 'w', encoding='utf-8') as f:
    f.write("=" * 100 + "\n")
    f.write("信号替换分析报告 (改进版)\n")
    f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("=" * 100 + "\n\n")
    
    f.write(f"源文件: {source_file}\n")
    f.write(f"映射文件: {mapping_file}\n")
    f.write(f"备份文件: {backup_file}\n\n")
    
    f.write(f"建立的映射关系总数: {len(signal_mapping)}\n")
    f.write(f"成功替换数: {len(successful)}\n")
    f.write(f"未找到映射数: {len(not_found)}\n\n")
    
    if successful:
        f.write("=" * 100 + "\n")
        f.write("成功替换的详情 (按Sheet分组):\n")
        f.write("=" * 100 + "\n")
        
        # 按Sheet分组
        successful_by_sheet = {}
        for r in successful:
            sheet = r['sheet']
            if sheet not in successful_by_sheet:
                successful_by_sheet[sheet] = []
            successful_by_sheet[sheet].append(r)
        
        for sheet_name in sorted(successful_by_sheet.keys()):
            replacements = successful_by_sheet[sheet_name]
            f.write(f"\nSheet: {sheet_name} ({len(replacements)} 项)\n")
            f.write("-" * 100 + "\n")
            for r in replacements:
                f.write(f"  行 {r['row']:5d}, 列 {r['col']:5d}: {r['old_signal']:60s} -> {r['new_signal']}\n")
    
    if not_found:
        f.write("\n" + "=" * 100 + "\n")
        f.write("未找到映射的详情 (按Sheet分组):\n")
        f.write("=" * 100 + "\n")
        
        # 按Sheet分组
        not_found_by_sheet = {}
        for r in not_found:
            sheet = r['sheet']
            if sheet not in not_found_by_sheet:
                not_found_by_sheet[sheet] = []
            not_found_by_sheet[sheet].append(r)
        
        for sheet_name in sorted(not_found_by_sheet.keys()):
            replacements = not_found_by_sheet[sheet_name]
            f.write(f"\nSheet: {sheet_name} ({len(replacements)} 项)\n")
            f.write("-" * 100 + "\n")
            
            # 按unique signal分组统计
            signal_counts = {}
            for r in replacements:
                sig = r['old_signal']
                if sig not in signal_counts:
                    signal_counts[sig] = 0
                signal_counts[sig] += 1
            
            for sig in sorted(signal_counts.keys()):
                count = signal_counts[sig]
                f.write(f"  信号 '{sig}' 出现 {count} 次\n")
                # 找出所有该信号出现的行
                for r in replacements:
                    if r['old_signal'] == sig:
                        f.write(f"    - 行 {r['row']:5d}, 列 {r['col']:5d}\n")
    
    f.write("\n" + "=" * 100 + "\n")
    f.write("原因分析:\n")
    f.write("=" * 100 + "\n\n")
    
    # 分析未找到的原因
    if not_found:
        f.write("未找到映射的原因分析:\n")
        
        # 统计各类未找到的信号
        has_tilde = 0
        has_html = 0
        has_table = 0
        has_image = 0
        
        for r in not_found:
            sig = r['old_signal']
            if '~' in sig:
                has_tilde += 1
            if 'Table' in sig or '|' in sig or '_x000D_' in sig:
                has_table += 1
            if 'Image' in sig or 'wiki=' in sig:
                has_image += 1
            if '<' in sig or '>' in sig or 'alt=' in sig:
                has_html += 1
        
        f.write(f"  1. 包含'~'符号的信号: {has_tilde} 项\n")
        f.write(f"     原因: Excel可能含有特殊编码或格式化标记\n\n")
        f.write(f"  2. 包含HTML/Table标记的内容: {has_table} 项\n")
        f.write(f"     原因: 单元格中可能包含表格或富文本内容\n\n")
        f.write(f"  3. 包含Image标记的内容: {has_image} 项\n")
        f.write(f"     原因: 单元格中可能包含图像或特殊对象\n\n")
        
        f.write(f"  4. 其他原因:\n")
        f.write(f"     - 映射文件中可能不存在该信号\n")
        f.write(f"     - 信号名拼写不匹配 (空格、大小写等)\n\n")

print(f"\n详细报告已保存: {report_file}\n")

# 生成简洁总结
summary_file = f"替换总结_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
with open(summary_file, 'w', encoding='utf-8') as f:
    f.write("信号替换总结\n")
    f.write("=" * 50 + "\n\n")
    f.write(f"源文件: {source_file}\n")
    f.write(f"备份文件: {backup_file}\n\n")
    f.write(f"成功替换: {len(successful)} 项\n")
    f.write(f"未找到:   {len(not_found)} 项\n")
    f.write(f"总计:     {len(successful) + len(not_found)} 项\n\n")
    
    f.write("建议:\n")
    f.write("1. 检查Excel文件中含有'~'的信号是否为格式化或特殊标记\n")
    f.write("2. 验证映射文件是否包含所有需要的信号\n")
    f.write("3. 检查System Requirements中是否有其他需要替换的信号\n")

print(f"总结已保存: {summary_file}")

print("\n" + "=" * 80)
print("任务完成!")
print("=" * 80)
